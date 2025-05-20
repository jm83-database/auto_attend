from flask import Flask, request, render_template, flash, redirect, url_for, send_from_directory
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import io
import csv
import threading
import time
import sys
import logging
from datetime import datetime, timedelta

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Azure에서 발생할 수 있는 오류를 로깅
try:
    app = Flask(__name__)
    app.secret_key = os.environ.get('SECRET_KEY', 'auto_attend_secret_key')

    # 파일 경로 상수 (상대 경로로 변경)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
    TEMPLATE_FOLDER = os.path.join(BASE_DIR, 'templates')
    RESULT_FOLDER = os.path.join(BASE_DIR, 'results')
    STATIC_RESULT_FOLDER = os.path.join(BASE_DIR, 'static', 'results')

    # 디버깅을 위한 경로 로깅
    logger.info(f"BASE_DIR: {BASE_DIR}")
    logger.info(f"UPLOAD_FOLDER: {UPLOAD_FOLDER}")
    logger.info(f"RESULT_FOLDER: {RESULT_FOLDER}")
    logger.info(f"STATIC_RESULT_FOLDER: {STATIC_RESULT_FOLDER}")

    # 정적 파일을 제공하기 위한 경로 설정
    app.static_folder = os.path.join(BASE_DIR, 'static')
    
    # 폴더 생성 시도
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        os.makedirs(RESULT_FOLDER, exist_ok=True)
        os.makedirs(STATIC_RESULT_FOLDER, exist_ok=True)
        logger.info("폴더 생성 성공")
    except Exception as e:
        logger.error(f"폴더 생성 중 오류 발생: {str(e)}")

except Exception as e:
    logger.error(f"앱 초기화 중 오류 발생: {str(e)}")
    raise

# 한국 시간대(KST)로 변환하는 함수
def get_korea_time():
    # UTC 시간에 9시간 추가
    return datetime.now() + timedelta(hours=9)

# 파일 삭제 함수 (1분 후)
def delete_file_after_delay(file_path, delay=60):
    def delete_task():
        time.sleep(delay)
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"파일 삭제 완료: {file_path}")
                
                # 정적 폴더의 파일도 삭제
                static_path = file_path.replace(RESULT_FOLDER, STATIC_RESULT_FOLDER)
                if os.path.exists(static_path):
                    os.remove(static_path)
                    logger.info(f"정적 파일 삭제 완료: {static_path}")
        except Exception as e:
            logger.error(f"파일 삭제 중 오류 발생: {str(e)}")
    
    # 새 스레드에서 삭제 작업 실행
    threading.Thread(target=delete_task).start()

# 참석보고서 데이터 처리 함수
def process_attendance_report(report_file):
    try:
        # UTF-16LE로 인코딩된 CSV 파일 읽기
        content = report_file.read()
        
        # 바이트 데이터를 UTF-16LE로 디코딩
        text = content.decode('utf-16le')
        
        # 텍스트를 라인별로 분리
        lines = text.splitlines()
        
        # 참가자 영역 찾기
        participant_start = -1
        for i, line in enumerate(lines):
            if line.startswith('2. 참가자'):
                participant_start = i
                break
        
        if participant_start == -1:
            logger.error("참가자 섹션을 찾을 수 없습니다.")
            return None
        
        # 헤더 및 데이터 행 구분
        header_line = lines[participant_start + 1]
        data_lines = []
        
        # 3. 모임 내 활동 전까지의 데이터 수집
        for i in range(participant_start + 2, len(lines)):
            if lines[i].startswith('3. 모임 내 활동'):
                break
            if lines[i].strip():  # 빈 줄 제외
                data_lines.append(lines[i])
        
        # 헤더 및 데이터 파싱
        headers = header_line.split('\t')
        
        participants = []
        for line in data_lines:
            if not line.strip():
                continue
            values = line.split('\t')
            if len(values) >= 4:  # 최소한 이름, 들어온 시간, 나간 시간, 참여 시간이 있어야 함
                participant = {
                    '이름': values[0],
                    '처음 들어온 시간': values[1],
                    '마지막 나간 시간': values[2],
                    '모임 참여 시간': values[3]
                }
                participants.append(participant)
        
        logger.info(f"참가자 {len(participants)}명의 데이터를 성공적으로 처리했습니다.")
        return participants
    except Exception as e:
        logger.error(f"참석보고서 처리 중 오류 발생: {str(e)}")
        raise

# 출석 데이터 처리 함수
def process_attendance_data(attendance_file):
    try:
        # CSV 파일 읽기
        df = pd.read_csv(attendance_file, encoding='utf-8')
        
        # 출석한 학생들만 필터링
        attendance_records = df[df['출석여부'] == '출석']
        
        # 이름과 출석 상태 추출
        attendance_dict = {}
        for _, row in attendance_records.iterrows():
            attendance_dict[row['이름']] = '출석'
        
        logger.info(f"출석 데이터 {len(attendance_dict)}명의 정보를 처리했습니다.")
        return attendance_dict
    except Exception as e:
        logger.error(f"출석 데이터 처리 중 오류 발생: {str(e)}")
        raise

# 엑셀 업데이트 함수
def update_excel(excel_file, attendance_dict, participants_list):
    try:
        # 엑셀 파일 로드
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['출결정보']
        
        # 열 인덱스 확인 (행 6에서 컬럼명 찾기)
        columns = {}
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=6, column=col).value
            if header in ['접속시작시간', '접속종료시간', '활용시간']:
                columns[header] = col
        
        # 중간출결 열 확인 (행 5에서 찾기)
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=5, column=col).value
            if header == '중간출결':
                columns['중간출결'] = col
                break
        
        # 이름 열 확인
        name_col = None
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=5, column=col).value
            if header == '성명':
                name_col = col
                break
        
        if not name_col or '중간출결' not in columns:
            logger.error("성명 또는 중간출결 열을 찾을 수 없습니다.")
            raise ValueError("성명 또는 중간출결 열을 찾을 수 없습니다.")
        
        # 학생 이름 검색 및 데이터 업데이트
        updates = {
            'attendance': 0,
            'login_time': 0,
            'logout_time': 0,
            'duration': 0
        }
        
        # 변경사항 추적을 위한 색상 설정 - #FFB366 (주황색)
        highlight_fill = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')
        
        # 시작 행 (헤더 다음 행부터)
        start_row = 7
        
        for row in range(start_row, sheet.max_row + 1):
            name = sheet.cell(row=row, column=name_col).value
            if not name:
                continue
            
            # 출석 정보 업데이트
            if name in attendance_dict:
                attendance_cell = sheet.cell(row=row, column=columns['중간출결'])
                if not attendance_cell.value or attendance_cell.value != 'O':
                    attendance_cell.value = 'O'
                    attendance_cell.fill = highlight_fill
                    updates['attendance'] += 1
            
            # 참석보고서 데이터 업데이트
            for participant in participants_list:
                if participant['이름'].strip() == name.strip():
                    # 접속시작시간 업데이트
                    if '접속시작시간' in columns:
                        login_cell = sheet.cell(row=row, column=columns['접속시작시간'])
                        login_time = participant['처음 들어온 시간']
                        if not login_cell.value or login_cell.value != login_time:
                            login_cell.value = login_time
                            login_cell.fill = highlight_fill
                            updates['login_time'] += 1
                    
                    # 접속종료시간 업데이트
                    if '접속종료시간' in columns:
                        logout_cell = sheet.cell(row=row, column=columns['접속종료시간'])
                        logout_time = participant['마지막 나간 시간']
                        if not logout_cell.value or logout_cell.value != logout_time:
                            logout_cell.value = logout_time
                            logout_cell.fill = highlight_fill
                            updates['logout_time'] += 1
                    
                    # 활용시간 업데이트
                    if '활용시간' in columns:
                        duration_cell = sheet.cell(row=row, column=columns['활용시간'])
                        duration = participant['모임 참여 시간']
                        if not duration_cell.value or duration_cell.value != duration:
                            duration_cell.value = duration
                            duration_cell.fill = highlight_fill
                            updates['duration'] += 1
                    
                    break
        
        # 한국 시간(KST)으로 타임스탬프 파일명 생성
        kst_now = get_korea_time()
        timestamp = kst_now.strftime("%Y%m%d_%H%M%S")
        result_filename = f'updated_{timestamp}_{os.path.basename(excel_file.filename)}'
        
        # 저장 경로 설정
        result_path = os.path.join(RESULT_FOLDER, result_filename)
        
        # 정적 폴더 경로 설정
        static_result_path = os.path.join(STATIC_RESULT_FOLDER, result_filename)
        
        # 업데이트된 엑셀 파일 저장 시도
        try:
            workbook.save(result_path)
            logger.info(f"결과 파일 저장 성공: {result_path}")
            workbook.save(static_result_path)
            logger.info(f"정적 파일 저장 성공: {static_result_path}")
        except Exception as e:
            logger.error(f"파일 저장 중 오류 발생: {str(e)}")
            raise
        
        return {
            'updates': updates,
            'result_path': result_path,
            'result_filename': result_filename,
            'full_result_path': result_path
        }
    except Exception as e:
        logger.error(f"엑셀 업데이트 중 오류 발생: {str(e)}")
        raise

@app.route('/', methods=['GET', 'POST'])
def index():
    try:
        logger.info("인덱스 페이지 접근")
        result = None
        
        if request.method == 'POST':
            logger.info("POST 요청 처리 시작")
            # 파일 업로드 검증
            if 'excel_file' not in request.files or 'attendance_file' not in request.files or 'report_file' not in request.files:
                flash('모든 필수 파일을 업로드해주세요.')
                return redirect(request.url)
            
            excel_file = request.files['excel_file']
            attendance_file = request.files['attendance_file']
            report_file = request.files['report_file']
            
            logger.info(f"업로드된 파일: {excel_file.filename}, {attendance_file.filename}, {report_file.filename}")
            
            if excel_file.filename == '' or attendance_file.filename == '' or report_file.filename == '':
                flash('모든 필수 파일을 선택해주세요.')
                return redirect(request.url)
            
            # 파일 처리
            try:
                # 출석 데이터 처리
                attendance_dict = process_attendance_data(attendance_file)
                
                # 참석보고서 처리
                participants_list = process_attendance_report(report_file)
                
                if not participants_list:
                    flash('참석보고서 파일을 처리하는 중 오류가 발생했습니다.')
                    return redirect(request.url)
                
                # 엑셀 파일 업데이트
                result = update_excel(excel_file, attendance_dict, participants_list)
                
                # 1분 후 파일 삭제 예약
                delete_file_after_delay(result['full_result_path'])
                
                flash('업데이트가 완료되었습니다. 변경된 셀은 주황색(#FFB366)으로 표시됩니다. 파일은 다운로드 후 1분 뒤에 자동으로 삭제됩니다.')
            
            except Exception as e:
                logger.error(f"파일 처리 중 오류 발생: {str(e)}")
                flash(f'오류 발생: {str(e)}')
                return redirect(request.url)
        
        return render_template('index.html', result=result)
    except Exception as e:
        logger.error(f"인덱스 페이지 처리 중 오류 발생: {str(e)}")
        flash(f'처리 중 오류가 발생했습니다: {str(e)}')
        return redirect('/')

@app.route('/download/<path:filename>')
def download_file(filename):
    try:
        logger.info(f"파일 다운로드 요청: {filename}")
        return send_from_directory(RESULT_FOLDER, filename, as_attachment=True)
    except Exception as e:
        logger.error(f"파일 다운로드 중 오류 발생: {str(e)}")
        flash(f'파일을 다운로드할 수 없습니다: {str(e)}')
        return redirect('/')

# Azure App Service 환경 변수 지원
port = int(os.environ.get('PORT', 5000))

# Flask 애플리케이션이 Azure WebApp에서 정상적으로 실행되는지 확인하기 위한 상태 체크 엔드포인트
@app.route('/health')
def health_check():
    # 현재 한국 시간을 확인하기 위한 정보 추가
    kst_now = get_korea_time()
    return f'OK - KST: {kst_now.strftime("%Y-%m-%d %H:%M:%S")}', 200

# Azure App Service에서 실행될 때 필요한 설정
if __name__ == '__main__':
    logger.info(f"Flask 앱 실행: 호스트=0.0.0.0, 포트={port}")
    app.run(host='0.0.0.0', port=port, debug=False)
